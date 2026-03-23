"""
Sistema de Registro de Asistencia con Buffer de Alta Concurrencia
=================================================================

Esta versión usa DuckDB como buffer en memoria para manejar
alta concurrencia (1000+ usuarios simultáneos).

Características:
- Escrituras instantáneas (<100ms)
- Sincronización automática con Google Sheets cada 60 segundos
- Panel de monitoreo de sincronización
- Recuperación automática de errores

Para usar este archivo:
1. pip install duckdb
2. Renombrar AsistenciaCurso.py a AsistenciaCurso_OLD.py
3. Renombrar este archivo a AsistenciaCurso.py
"""

import streamlit as st
import pandas as pd
import time
import requests
from datetime import datetime, date
from rut_chile import rut_chile
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Importar el sistema de buffer
from db_buffer import get_buffer

# Configuración básica
st.set_page_config(page_title="Registro de Asistencia", layout="wide", initial_sidebar_state="collapsed")

# Constantes
SECRET_PASSWORD = st.secrets["SECRET_PASSWORD"]
API_URL = st.secrets["API_URL"]
API_KEY = st.secrets["API_KEY"]

# ==================== FUNCIONES DE API ====================

# Función para obtener datos de configuración de cursos
@st.cache_data(ttl=300)  # Cache por 5 minutos
def get_config_data():
    try:
        response = requests.get(f"{API_URL}?action=getConfig&key={API_KEY}")
        data = response.json()

        if data['success']:
            df = pd.DataFrame(data['cursos'])
            if not df.empty:
                # Convertir columnas de fecha a datetime (detectando formato automáticamente)
                date_cols = ['fecha_inicio', 'fecha_fin', 'fecha_jornada', 'fecha_sesion_1', 'fecha_sesion_2', 'fecha_sesion_3']
                for col in date_cols:
                    if col in df.columns:
                        parsed = pd.to_datetime(df[col], dayfirst=True, errors='coerce')
                        if parsed.dt.tz is not None:
                            parsed = parsed.dt.tz_convert(None)
                        df[col] = parsed.dt.normalize()

                df['cupo_maximo'] = pd.to_numeric(df['cupo_maximo'], errors='coerce')
            return df
        else:
            st.error(f"Error al obtener configuración: {data.get('error', 'Error desconocido')}")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error al conectar con la API: {str(e)}")
        return pd.DataFrame()

# Función para obtener registros de inscripción
@st.cache_data(ttl=180)  # Cache por 3 minutos
def get_registros_data():
    try:
        response = requests.get(f"{API_URL}?action=getRegistros&key={API_KEY}")
        data = response.json()

        if data['success']:
            return pd.DataFrame(data['registros'])
        else:
            st.error(f"Error al obtener registros: {data.get('error', 'Error desconocido')}")
            return pd.DataFrame()
    except Exception as e:
        st.error(f"Error al conectar con la API: {str(e)}")
        return pd.DataFrame()

# ==================== FUNCIONES DE BUFFER ====================

def guardar_asistencia_buffer(curso_id, rut, sesion):
    """
    Guarda asistencia en el buffer local (instantáneo).

    Args:
        curso_id: ID del curso
        rut: RUT del participante
        sesion: Número de sesión

    Returns:
        dict: Resultado de la operación
    """
    buffer = get_buffer()

    # Verificar si ya existe
    if buffer.verificar_asistencia(curso_id, rut, sesion):
        return {
            'success': False,
            'message': 'Ya existe un registro de asistencia para este participante en esta sesión'
        }

    # Marcar asistencia en buffer (ultra rápido)
    resultado = buffer.marcar_asistencia(
        curso_id=curso_id,
        rut=rut,
        sesion=sesion,
        estado='presente',
        metodo='streamlit_buffer'
    )

    return resultado

def get_asistencias_from_buffer(curso_id=None, sesion=None):
    """
    Obtiene asistencias desde el buffer local (instantáneo).

    Args:
        curso_id: ID del curso (opcional)
        sesion: Número de sesión (opcional)

    Returns:
        pd.DataFrame: DataFrame con asistencias
    """
    buffer = get_buffer()

    if curso_id and sesion:
        return buffer.get_asistencias_curso(curso_id, sesion)
    elif curso_id:
        return buffer.get_asistencias_curso(curso_id)
    else:
        # Obtener todas las asistencias
        return buffer.conn.execute("SELECT * FROM asistencias_buffer").df()

# ==================== FUNCIONES AUXILIARES ====================

def get_cursos_con_sesion_hoy(df_cursos):
    """
    Filtra cursos que tienen sesión hoy y devuelve DataFrame con información adicional.

    Args:
        df_cursos: DataFrame con configuración de cursos

    Returns:
        pd.DataFrame: Cursos con sesión hoy incluyendo columna 'sesion_hoy'
    """
    if df_cursos.empty:
        return pd.DataFrame()

    hoy_d = pd.Timestamp.now().date()
    cursos_hoy = []

    for _, curso in df_cursos.iterrows():
        encontrado = False

        # 1. Verificar fecha_jornada (sesión única explícita)
        if 'fecha_jornada' in curso and pd.notna(curso['fecha_jornada']):
            if pd.to_datetime(curso['fecha_jornada']).date() == hoy_d:
                curso_dict = curso.to_dict()
                curso_dict['sesion_hoy'] = 1
                curso_dict['fecha_sesion_hoy'] = curso['fecha_jornada']
                cursos_hoy.append(curso_dict)
                encontrado = True

        # 2. Verificar fecha_sesion_1/2/3 (cursos multi-sesión)
        if not encontrado:
            for sesion_num in [1, 2, 3]:
                fecha_col = f'fecha_sesion_{sesion_num}'
                if fecha_col in curso and pd.notna(curso[fecha_col]):
                    if pd.to_datetime(curso[fecha_col]).date() == hoy_d:
                        curso_dict = curso.to_dict()
                        curso_dict['sesion_hoy'] = sesion_num
                        curso_dict['fecha_sesion_hoy'] = curso[fecha_col]
                        cursos_hoy.append(curso_dict)
                        encontrado = True
                        break

        # 3. Fallback: hoy está dentro del rango fecha_inicio – fecha_fin
        if not encontrado:
            inicio = curso.get('fecha_inicio')
            fin = curso.get('fecha_fin')
            if pd.notna(inicio) and pd.notna(fin):
                inicio_d = pd.to_datetime(inicio, dayfirst=True).date()
                fin_d = pd.to_datetime(fin, dayfirst=True).date()
                if inicio_d <= hoy_d <= fin_d:
                    curso_dict = curso.to_dict()
                    curso_dict['sesion_hoy'] = 1
                    curso_dict['fecha_sesion_hoy'] = inicio
                    cursos_hoy.append(curso_dict)

    if cursos_hoy:
        return pd.DataFrame(cursos_hoy)
    else:
        return pd.DataFrame()

def validar_participante_inscrito(rut, curso_id, df_registros):
    """
    Verifica si un participante está inscrito en un curso.

    Args:
        rut: RUT del participante
        curso_id: ID del curso
        df_registros: DataFrame con registros de inscripciones

    Returns:
        tuple: (bool, dict) - (está_inscrito, datos_participante)
    """
    if df_registros.empty:
        return False, None

    # Buscar participante — comparación case-insensitive (ej: 12345678-k == 12345678-K)
    rut_norm = str(rut).upper().strip()
    participante = df_registros[
        (df_registros['rut'].astype(str).str.upper().str.strip() == rut_norm) &
        (df_registros['curso_id'] == curso_id)
    ]

    if not participante.empty:
        return True, participante.iloc[0].to_dict()
    else:
        return False, None

# ==================== GENERACIÓN DE REPORTES EXCEL ====================

def _split_rut(rut_str):
    partes = str(rut_str).strip().upper().split('-')
    return (partes[0].replace('.', ''), partes[1]) if len(partes) == 2 else (rut_str, '')

def _sexo_codigo(sexo):
    return 1 if str(sexo).upper() == 'HOMBRE' else 2

def _nac_codigo(nac):
    return 1 if str(nac).upper() == 'CHILENO' else 2

_ROL_MK = {'PROFESIONAL SST': 1, 'TRABAJADOR': 2, 'MIEMBRO DE COMITÉ PARITARIO': 3,
            'MIEMBRO COMITE PARITARIO': 3, 'MONITOR O DELEGADO': 4, 'DIRIGENTE SINDICAL': 5}

def generar_excel_ist(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Campos inscripción ISTeduca"
    headers = ["RUT trabajador (Sin puntos ni dv)", "DV", "Nombres", "Apellidos (ambos)",
               "Email (individual)", "Género", "Rol trabajador", "Región", "Comuna",
               "Rut empresa (Sin puntos, con guión)", "Razón social"]
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style='thin', color="AAAAAA")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = brd
    ws.row_dimensions[1].height = 30
    for c, w in enumerate([22,6,25,30,30,10,25,30,25,28,35], 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    df_ = Font(name="Arial", size=10)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        rb, dv = _split_rut(getattr(row, 'rut', ''))
        ap = f"{getattr(row,'apellido_paterno','')} {getattr(row,'apellido_materno','')}".strip()
        for c, v in enumerate([rb, dv, getattr(row,'nombres',''), ap,
            getattr(row,'email',''), getattr(row,'sexo',''), getattr(row,'rol',''),
            getattr(row,'region',''), getattr(row,'comuna',''),
            getattr(row,'rut_empresa',''), getattr(row,'razon_social','')], 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.font = df_; cell.border = brd
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def generar_excel_mk(df):
    wb = Workbook()
    ws = wb.active; ws.title = "Datos"
    headers = ["Rut","Nombres","Apellido Paterno","Apellido Materno",
               "Sexo","Nacionalidad","Rol Trabajador","Otro Rol"]
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="2E75B6")
    thin = Side(style='thin', color="AAAAAA")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill
        cell.alignment = Alignment(horizontal="center"); cell.border = brd
    for c, w in enumerate([18,25,25,25,8,14,16,25], 1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w
    df_ = Font(name="Arial", size=10)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        rol = str(getattr(row,'rol','')).upper()
        rc = _ROL_MK.get(rol, 2)
        otro = rol if rol not in _ROL_MK else ''
        for c, v in enumerate([getattr(row,'rut',''), getattr(row,'nombres',''),
            getattr(row,'apellido_paterno',''), getattr(row,'apellido_materno',''),
            _sexo_codigo(getattr(row,'sexo','')), _nac_codigo(getattr(row,'nacionalidad','')),
            rc, otro], 1):
            cell = ws.cell(row=ri, column=c, value=v)
            cell.font = df_; cell.border = brd
    for sh, rows in [("Parametros",[("Descripcion","Valor"),("Largo máximo Rut",15),
                      ("Largo máximo nombres",50),("Largo máximo apellido paterno",50),
                      ("Largo máximo apellido materno",50)]),
                     ("MaeSexo",[("Codigo","Valor"),(1,"Hombre"),(2,"Mujer")]),
                     ("MaeNacionalidad",[("Codigo","Valor"),(1,"Chileno"),(2,"Extranjero")]),
                     ("MaeRolTrabajador",[("Codigo","Valor"),(1,"Profesional SST"),(2,"Trabajador"),
                      (3,"Miembro Comité Paritario"),(4,"Monitor o Delegado"),(5,"Dirigente Sindical")])]:
        ws2 = wb.create_sheet(sh)
        for r in rows: ws2.append(r)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ==================== INTERFAZ PRINCIPAL ====================

def main():
    st.title("📋 Registro de Asistencia - Jornada de Difusión sobre el Nuevo Protocolo de Ruido ISP (Res. Ex. Nº 5.921) - Empresas Adherentes de IST")

    # Obtener instancia del buffer
    buffer = get_buffer()

    # ==================== SIDEBAR CON AUTENTICACIÓN ====================

    st.sidebar.title("🔐 Panel de Control")

    # Autenticación admin
    password = st.sidebar.text_input("Contraseña Admin", type="password", key="admin_password")
    admin_mode = password == SECRET_PASSWORD

    if admin_mode:
        st.sidebar.success("✅ Acceso Admin")

        # Mostrar estadísticas del buffer (solo para admin)
        st.sidebar.divider()
        st.sidebar.subheader("📊 Estado del Buffer")
        stats = buffer.get_estadisticas()

        col1, col2 = st.sidebar.columns(2)
        with col1:
            st.metric("Total", stats['total'])
            st.metric("Sincronizadas", stats['sincronizadas'])
        with col2:
            st.metric("Pendientes", stats['pendientes'])
            st.metric("Fallidas", stats['fallidas'])

        st.sidebar.divider()

        # Botones de control (solo para admin)
        if st.sidebar.button("🔄 Sincronizar Ahora"):
            with st.spinner("Sincronizando con Google Sheets..."):
                resultado = buffer.sincronizar(batch_size=300)  # Aumentado a 300

            st.sidebar.success(f"✅ Sincronizados: {resultado['sincronizados']}")
            if resultado['fallidos'] > 0:
                st.sidebar.warning(f"⚠️ Fallidos: {resultado['fallidos']}")

        if st.sidebar.button("🗑️ Limpiar Sincronizados"):
            eliminados = buffer.limpiar_sincronizados(dias=0)
            st.sidebar.success(f"✅ Eliminados: {eliminados} registros")

        if st.sidebar.button("🚨 Borrar Todo el Buffer", type="primary"):
            buffer.conn.execute("DELETE FROM asistencias_buffer")
            buffer.hydrate_from_sheets()
            st.sidebar.success("✅ Buffer vaciado y recargado desde Sheets")
            st.rerun()
    else:
        if password:
            st.sidebar.error("❌ Contraseña incorrecta")

    st.sidebar.divider()

    # Botón para limpiar cache (útil si se actualizaron datos en Sheets)
    if st.sidebar.button("🔄 Actualizar datos"):
        st.cache_data.clear()
        st.rerun()

    # ==================== MODO PARTICIPANTE (SIN PASSWORD) ====================

    if not admin_mode:
        st.info("👤 **Modo Participante:** Marca tu asistencia ingresando tu RUT")

        # Obtener cursos con sesión hoy
        df_cursos = get_config_data()
        df_cursos_hoy = get_cursos_con_sesion_hoy(df_cursos)

        if df_cursos_hoy.empty:
            st.warning("⚠️ No hay cursos con sesión programada para hoy.")
            st.stop()

        # Mostrar cursos disponibles
        st.subheader("📅 Cursos con Sesión Hoy")

        @st.fragment
        def formulario_asistencia(curso_id, sesion_hoy, region, fecha_str):
            with st.expander(f"📚 {curso_id}", expanded=True):
                st.write(f"**Región:** {region}")
                st.write(f"**Fecha:** {fecha_str}")

                with st.form(key=f"form_{curso_id}_{sesion_hoy}", clear_on_submit=True):
                    rut_input = st.text_input(
                        "Ingresa tu RUT (sin puntos, con guión)",
                        placeholder="12345678-9"
                    )

                    submit = st.form_submit_button("✅ Marcar Asistencia")

                    if submit and rut_input:
                        rut_input = rut_input.strip().upper()

                        if not rut_chile.is_valid_rut(rut_input):
                            st.error("❌ RUT inválido. Verifica el formato.")
                        else:
                            df_registros = get_registros_data()
                            esta_inscrito, datos = validar_participante_inscrito(
                                rut_input, curso_id, df_registros
                            )

                            if not esta_inscrito:
                                st.error("❌ No estás inscrito en este curso. Contacta al administrador.")
                            else:
                                resultado = guardar_asistencia_buffer(
                                    curso_id=curso_id,
                                    rut=rut_input,
                                    sesion=sesion_hoy
                                )

                                if resultado['success']:
                                    nombre_completo = f"{datos.get('nombres', '')} {datos.get('apellido_paterno', '')}".strip() or rut_input
                                    st.success(f"✅ ¡Asistencia registrada para {nombre_completo}!")
                                    st.info("🎉 Ya puedes cerrar esta pestaña.")
                                    st.balloons()
                                else:
                                    st.warning(f"ℹ️ {resultado['message']}")

        for _, curso in df_cursos_hoy.iterrows():
            formulario_asistencia(
                curso_id=curso['curso_id'],
                sesion_hoy=curso['sesion_hoy'],
                region=curso.get('region', 'N/A'),
                fecha_str=curso['fecha_sesion_hoy'].strftime('%d-%m-%Y')
            )

        st.stop()

    # ==================== MODO ADMIN ====================

    if admin_mode:
        # Tabs para diferentes funciones
        tab1, tab2, tab3 = st.tabs(["📝 Gestionar Asistencia", "📊 Ver Asistencias", "🔧 Mantenimiento"])

        # TAB 1: Gestionar Asistencia Manual
        with tab1:
            st.subheader("📝 Registro Manual de Asistencia")

            df_cursos = get_config_data()

            if df_cursos.empty:
                st.warning("⚠️ No hay cursos disponibles")
            else:
                # Seleccionar curso
                curso_ids = df_cursos['curso_id'].tolist()
                curso_seleccionado = st.selectbox("Selecciona un curso", curso_ids)

                # Obtener info del curso
                curso = df_cursos[df_cursos['curso_id'] == curso_seleccionado].iloc[0]

                # Mostrar sesiones disponibles
                sesiones = []
                for i in [1, 2, 3]:
                    if f'fecha_sesion_{i}' in curso and pd.notna(curso[f'fecha_sesion_{i}']):
                        sesiones.append(i)
                # Fallback: fecha_jornada o fecha_inicio cuentan como sesión 1
                if not sesiones:
                    if ('fecha_jornada' in curso and pd.notna(curso['fecha_jornada'])) or \
                       ('fecha_inicio' in curso and pd.notna(curso['fecha_inicio'])):
                        sesiones = [1]

                if not sesiones:
                    st.warning("⚠️ Este curso no tiene sesiones configuradas")
                else:
                    sesion_seleccionada = st.selectbox("Selecciona sesión", sesiones)

                    # Formulario de registro
                    with st.form("form_admin"):
                        col1, col2 = st.columns(2)

                        with col1:
                            rut = st.text_input("RUT", placeholder="12345678-9")

                        with col2:
                            estado = st.selectbox("Estado", ["presente", "ausente", "justificado"])

                        submit = st.form_submit_button("💾 Registrar")

                        if submit and rut:
                            if not rut_chile.is_valid_rut(rut):
                                st.error("❌ RUT inválido")
                            else:
                                # Verificar inscripción
                                df_registros = get_registros_data()
                                esta_inscrito, datos = validar_participante_inscrito(
                                    rut, curso_seleccionado, df_registros
                                )

                                if not esta_inscrito:
                                    st.error("❌ Participante no inscrito en este curso")
                                else:
                                    # Marcar en buffer
                                    resultado = buffer.marcar_asistencia(
                                        curso_id=curso_seleccionado,
                                        rut=rut,
                                        sesion=sesion_seleccionada,
                                        estado=estado,
                                        metodo='admin_manual'
                                    )

                                    if resultado['success']:
                                        nombre_completo = f"{datos.get('nombres', '')} {datos.get('apellido_paterno', '')}".strip() or rut
                                        st.success(f"✅ Asistencia registrada para {nombre_completo}")
                                    else:
                                        st.error(f"❌ {resultado['message']}")

        # TAB 2: Ver Asistencias
        with tab2:
            st.subheader("📊 Visualizar Asistencias")

            df_cursos = get_config_data()

            if not df_cursos.empty:
                curso_ids = df_cursos['curso_id'].tolist()
                curso_ver = st.selectbox("Curso", curso_ids, key="ver_curso")

                sesion_ver = st.selectbox("Sesión", [1, 2, 3], key="ver_sesion")

                # Obtener asistencias desde el buffer
                df_asist = get_asistencias_from_buffer(curso_ver, sesion_ver)

                if not df_asist.empty:
                    st.write(f"**Total registros:** {len(df_asist)}")

                    # Mostrar estado de sincronización
                    sincronizadas = df_asist['sincronizado'].sum()
                    pendientes = len(df_asist) - sincronizadas

                    col1, col2 = st.columns(2)
                    with col1:
                        st.metric("✅ Sincronizadas", sincronizadas)
                    with col2:
                        st.metric("⏳ Pendientes", pendientes)

                    # Mostrar tabla
                    st.dataframe(
                        df_asist[['rut', 'estado', 'fecha_registro', 'sincronizado', 'intentos_sync']],
                        use_container_width=True
                    )

                    # Botón para exportar CSV
                    csv = df_asist.to_csv(index=False).encode('utf-8')
                    st.download_button(
                        "📥 Descargar CSV",
                        csv,
                        f"asistencias_{curso_ver}_sesion_{sesion_ver}.csv",
                        "text/csv"
                    )

                    # Descargar Reportes Excel
                    st.divider()
                    st.subheader("📥 Descargar Reportes")
                    df_reg_rep = get_registros_data()
                    if not df_reg_rep.empty and 'rut' in df_reg_rep.columns and 'curso_id' in df_reg_rep.columns:
                        ruts_asist = df_asist['rut'].str.upper().str.strip().unique()
                        df_reg_c = df_reg_rep[df_reg_rep['curso_id'] == curso_ver].copy()
                        df_reg_c['rut_norm'] = df_reg_c['rut'].astype(str).str.upper().str.strip()
                        df_asistentes = df_reg_c[df_reg_c['rut_norm'].isin(ruts_asist)].copy()
                        if df_asistentes.empty:
                            st.info("ℹ️ No hay asistentes con datos de inscripción para exportar.")
                        else:
                            col_r1, col_r2 = st.columns(2)
                            with col_r1:
                                st.markdown("**Formato IST Educa**")
                                st.download_button(
                                    label="📥 Descargar IST Educa (.xlsx)",
                                    data=generar_excel_ist(df_asistentes),
                                    file_name=f"IST_{curso_ver}_s{sesion_ver}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                            with col_r2:
                                st.markdown("**Formato MK Capacitaciones**")
                                st.download_button(
                                    label="📥 Descargar MK Capacitaciones (.xlsx)",
                                    data=generar_excel_mk(df_asistentes),
                                    file_name=f"MK_{curso_ver}_s{sesion_ver}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                else:
                    st.info("ℹ️ No hay asistencias registradas para este curso y sesión")

        # TAB 3: Mantenimiento
        with tab3:
            st.subheader("🔧 Mantenimiento del Sistema")

            st.write("### Sincronización Manual")
            batch_size = st.number_input("Tamaño del lote", min_value=10, max_value=200, value=50)

            if st.button("🚀 Sincronizar Lote Completo"):
                with st.spinner("Sincronizando..."):
                    resultado = buffer.sincronizar(batch_size=batch_size)

                st.write("**Resultado:**")
                st.json(resultado)

            st.divider()

            st.write("### Limpieza de Registros")
            dias = st.number_input("Mantener últimos N días", min_value=1, max_value=30, value=7)

            if st.button("🗑️ Limpiar Registros Antiguos"):
                eliminados = buffer.limpiar_sincronizados(dias=dias)
                st.success(f"✅ Eliminados {eliminados} registros antiguos")


if __name__ == "__main__":
    main()
